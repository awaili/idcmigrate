"""RVTools adapter.

RVTools exports vSphere inventory as multi-tab CSV/xlsx. We read the three
tabs that matter for migration sizing: ``vInfo`` (VM summary), ``vDisk``
(per-disk capacity), ``vNetwork`` (NICs/IPs). Headers are matched flexibly
(case-insensitive, dots/parens/spaces normalized) so real RVTools exports
work without adjustment.

This adapter is always file-based (RVTools is an on-prem vSphere export,
there is no online API). It can ingest .csv; .xlsx needs ``openpyxl``.
"""
from __future__ import annotations

import csv
import os
import re
from typing import Any, Dict, List

from ...config import Settings
from ..models import RawAsset, SOURCE_RVTOOLS
from .base import Adapter, IngestResult, register


def _norm(s: str) -> str:
    return re.sub(r"[^a-z0-9]", "", (s or "").lower())


class _Row:
    """Dict with normalized-key lookup so we can match RVTools headers."""
    def __init__(self, raw: Dict[str, Any]):
        self.raw = raw
        self._map = {_norm(k): v for k, v in raw.items() if k}

    def get(self, *keys: str) -> str:
        for k in keys:
            v = self._map.get(_norm(k))
            if v not in (None, ""):
                return str(v)
        return ""

    def all(self) -> Dict[str, Any]:
        return {k: v for k, v in self.raw.items() if k}


@register
class RVToolsAdapter(Adapter):
    source = SOURCE_RVTOOLS

    def fetch(self, settings: Settings) -> IngestResult:
        path = settings.rvtools_path
        if not path or not os.path.exists(path):
            return IngestResult(assets=[], mode="fixture",
                                error=f"rvtools file not found: {path}")
        # RVTools is always file-based (vSphere export). Only block the BUNDLED
        # fixture when the fallback is disabled — an operator-uploaded real
        # file (outside the fixtures dir) is always honored.
        if not settings.allow_fixture_fallback:
            from ...config import FIXTURES
            try:
                if os.path.commonpath([os.path.abspath(path),
                                       str(FIXTURES)]) == str(FIXTURES):
                    return self._fixture_disabled(settings)
            except ValueError:
                pass   # different drives — not under fixtures, allow it
        if path.lower().endswith((".xlsx", ".xls")):
            return self._read_xlsx(path)
        return self._read_csv(path)

    # -- csv ---------------------------------------------------------------
    def _read_csv(self, vinfo_path: str) -> IngestResult:
        d = os.path.dirname(vinfo_path)
        vinfo_rows = self._load_csv(vinfo_path)
        vdisk_rows = self._maybe_load(d, "rvtools_vDisk.csv")
        vnet_rows = self._maybe_load(d, "rvtools_vNetwork.csv")

        disks_by_vm: Dict[str, List[Dict[str, Any]]] = {}
        for r in vdisk_rows:
            row = _Row(r)
            vm = row.get("VM", "VM Name")
            disks_by_vm.setdefault(vm, []).append({
                "name": row.get("Disk", "Disk Label"),
                "size_gb": _to_int(row.get("Capacity (GB)", "Provisioned (GB)", "Capacity GB")),
                "kind": row.get("Disk Type", "Type"),
                "fs": row.get("Path"),
            })

        net_by_vm: Dict[str, List[Dict[str, Any]]] = {}
        for r in vnet_rows:
            row = _Row(r)
            vm = row.get("VM", "VM Name")
            net_by_vm.setdefault(vm, []).append({
                "network": row.get("Network", "Network Name"),
                "ip": row.get("IP Address", "IPv4 Address"),
                "mac": row.get("MAC Address", "MAC"),
            })

        assets: List[RawAsset] = []
        for r in vinfo_rows:
            row = _Row(r)
            vm = row.get("VM", "VM Name", "Name")
            if not vm:
                continue
            ip = row.get("IP Address #1", "IP Address", "IPv4 Address")
            os_full = row.get("OS according to the configuration file", "OS", "Guest OS")
            attrs: Dict[str, Any] = {
                "power_state": row.get("PowerState", "Power State"),
                "os": os_full,
                "cpus": _to_int(row.get("CPUs", "CPU", "vCPU")),
                "memory_mb": _to_int(row.get("Memory", "Memory (MB)")),
                "provisioned_gb": _to_float(row.get("Provisioned (GB)", "Provisioned GB")),
                "used_gb": _to_float(row.get("Used (GB)", "Used GB")),
                "datacenter": row.get("Datacenter"),
                "cluster": row.get("Cluster"),
                "host": row.get("Host", "ESX Host"),
                "folder": row.get("Folder"),
                "disks": disks_by_vm.get(vm, []),
                "networks": net_by_vm.get(vm, []),
            }
            assets.append(RawAsset(
                source=SOURCE_RVTOOLS, source_id=vm,
                hostname=vm, fqdn="", ip=ip, attrs=attrs,
            ))
        return IngestResult(assets=assets, mode="fixture")

    def _load_csv(self, path: str) -> List[Dict[str, Any]]:
        with open(path, newline="", encoding="utf-8-sig") as f:
            return list(csv.DictReader(f))

    def _maybe_load(self, dirname: str, name: str) -> List[Dict[str, Any]]:
        p = os.path.join(dirname, name)
        return self._load_csv(p) if os.path.exists(p) else []

    # -- xlsx (optional) --------------------------------------------------
    def _read_xlsx(self, path: str) -> IngestResult:
        try:
            from openpyxl import load_workbook  # type: ignore
        except ImportError:
            return IngestResult(assets=[], mode="fixture",
                                error="openpyxl required to read .xlsx; pip install openpyxl")
        wb = load_workbook(path, read_only=True, data_only=True)
        out: Dict[str, List[Dict[str, Any]]] = {}
        for ws in wb.worksheets:
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue
            header = [str(h) if h is not None else "" for h in rows[0]]
            name = ws.title.strip().lower()
            out[name] = [dict(zip(header, r)) for r in rows[1:]]
        vinfo = out.get("vinfo", [])
        if not vinfo:
            return IngestResult(assets=[], mode="fixture",
                                error="no vInfo tab in xlsx")
        # build the same disk/network maps as the csv path from the vDisk /
        # vNetwork tabs (already loaded into `out`), so xlsx exports don't
        # silently drop per-disk + NIC inventory.
        disks_by_vm: Dict[str, List[Dict[str, Any]]] = {}
        for r in out.get("vdisk", []):
            row = _Row(r)
            vm = row.get("VM", "VM Name")
            disks_by_vm.setdefault(vm, []).append({
                "name": row.get("Disk", "Disk Label"),
                "size_gb": _to_int(row.get("Capacity (GB)", "Provisioned (GB)", "Capacity GB")),
                "kind": row.get("Disk Type", "Type"),
                "fs": row.get("Path"),
            })
        net_by_vm: Dict[str, List[Dict[str, Any]]] = {}
        for r in out.get("vnetwork", []):
            row = _Row(r)
            vm = row.get("VM", "VM Name")
            net_by_vm.setdefault(vm, []).append({
                "network": row.get("Network", "Network Name"),
                "ip": row.get("IP Address", "IPv4 Address"),
                "mac": row.get("MAC Address", "MAC"),
            })
        # normalize into the same shape as csv path
        assets: List[RawAsset] = []
        for r in vinfo:
            row = _Row(r)
            vm = row.get("VM", "Name")
            if not vm:
                continue
            assets.append(RawAsset(
                source=SOURCE_RVTOOLS, source_id=vm, hostname=vm,
                fqdn="", ip=row.get("IP Address #1", "IP Address"),
                attrs={"os": row.get("OS according to the configuration file", "OS"),
                       "cpus": _to_int(row.get("CPUs")),
                       "memory_mb": _to_int(row.get("Memory")),
                       "provisioned_gb": _to_float(row.get("Provisioned (GB)")),
                       "datacenter": row.get("Datacenter"), "cluster": row.get("Cluster"),
                       "folder": row.get("Folder"),
                       "disks": disks_by_vm.get(vm, []),
                       "networks": net_by_vm.get(vm, [])},
            ))
        return IngestResult(assets=assets, mode="fixture")


def _to_int(v: str) -> int:
    try:
        return int(float(v or 0))
    except (TypeError, ValueError):
        return 0


def _to_float(v: str) -> float:
    try:
        return float(v or 0)
    except (TypeError, ValueError):
        return 0.0